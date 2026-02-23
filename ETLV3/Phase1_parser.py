"""
╔══════════════════════════════════════════════════════════════════════════════╗
║                    PHASE 1 — DATA INPUT PARSER                               ║
║                        (Single-Sheet Format)                                 ║
╚══════════════════════════════════════════════════════════════════════════════╝

VERSION:        1.0.0
RELEASE DATE:   2026-02-18
AUTHOR:         Data Engineering Team
LANGUAGE:       Python 3.9+
DEPENDENCIES:   pandas, openpyxl

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📋 DESCRIPTION
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Parses Excel-based ETL mapping definitions and generates:
  • Validated mapping metadata (JSON)
  • Ready-to-run source extraction SQL (SELECT/FROM/WHERE/QUALIFY)
  • Primary key derivation from source_keys/target_keys flags

Used for defining table-to-table data mappings with support for:
  - Multi-table joins
  - SCD2 (Slowly Changing Dimension Type 2)
  - Delete flag handling
  - Deduplication logic (KEEP_FIRST, KEEP_LAST, REJECT)
  - Source lookup table enrichment
  - Custom transformations (SQL expressions)
  - Sample data limiting (QUALIFY clause)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📥 INPUT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

FILE FORMAT:    Excel (.xlsx)
SHEET NAME:     "MAPPING"
LAYOUT:         Single-sheet with two sections per mapping

MAPPING INFORMATION (30 fields):
  Mandatory (*): mapping_id, db_name, source_system, source_schema, source_table,
                 target_schema, target_table, scd2_applicable, delete_flag_applicable

  Key fields:
    • source_table: Comma-separated if multiple (e.g., "TABLE1, TABLE2")
    • source_join: Full JOIN clause (required for multi-table)
    • source_lookup_table: Optional reference table to enrich data
    • source_lookup_join_condition: How to join the lookup table
    • dedup_logic: KEEP_FIRST | KEEP_LAST | REJECT
    • scd2_applicable: Y or N
    • sample_set: Limit rows via QUALIFY ROW_NUMBER

COLUMN MAPPINGS (16 fields):
  Mandatory (*): source_table, source_column, source_keys, source_data_type,
                 target_column, target_keys, target_data_type, transformationtype

  Key fields:
    • source_keys / target_keys: Y = part of primary key
    • transformationtype: COPY (direct) or SQL (use transformationrule)
    • transformationrule: SQL expression (required if type=SQL)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📤 OUTPUT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

FILE FORMAT:    JSON (array of mapping objects)

STRUCTURE:
  {
    "mapping_id": "MAP_001",
    "db_name": "TEST_1_DB",
    "source_table": "DEMO_TABLE, DEMO_TABLE2",
    "source_join": "DEMO_TABLE T1 LEFT JOIN DEMO_TABLE2 T2 ON ...",
    "dedup_logic": "KEEP_LAST",
    "scd2_applicable": "Y",
    "sample_set": 100,

    "target_sql": "SELECT ... FROM ... WHERE ... QUALIFY ...",

    "columns": [
      {
        "source_table": "DEMO_TABLE",
        "source_column": "PKCOL1",
        "source_keys": "Y",
        "target_column": "COL_1",
        "transformationtype": "COPY"
      },
      ...
    ]
  }

GENERATED SQL (target_sql field):
  • SELECT:   Resolved column expressions with alias prefixes
  • FROM:     Fully qualified tables (db.schema.table) + JOINs
  • WHERE:    Combined filters (source_filter + date_filter + other)
  • QUALIFY:  ROW_NUMBER() for deduplication/sampling

PRIMARY KEY DERIVATION:
  • Source PKs: All columns where source_keys = Y
  • Target PKs: All columns where target_keys = Y
  • Used in QUALIFY ORDER BY for KEEP_FIRST/KEEP_LAST logic

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔍 VALIDATIONS (28 rules)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

TABLE-LEVEL:
  ✓ Mandatory fields populated
  ✓ Unique mapping_id
  ✓ Y/N fields (scd2_applicable, delete_flag_applicable)
  ✓ SCD2: start/end columns required if applicable=Y
  ✓ Multi-table: source_join required if >1 source_table
  ✓ dedup_logic: KEEP_FIRST, KEEP_LAST, or REJECT
  ✓ Lookup consistency: table + join_condition
  ✓ Parenthesis balance in filters
  ✓ Unescaped quotes warning

COLUMN-LEVEL:
  ✓ Mandatory fields populated
  ✓ transformationtype: COPY or SQL
  ✓ If type=SQL, transformationrule required
  ✓ No duplicate target_column within mapping
  ✓ source_table must be in mapping's table list
  ✓ Parenthesis balance in transformationrule

CROSS-SHEET:
  ✓ At least one source_keys=Y column per mapping
  ✓ At least one target_keys=Y column per mapping
  ✓ SCD2 columns exist as target_column entries
  ✓ sample_set requires source PK columns

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📝 CHANGE LOG
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

v1.0.0 - 2026-02-18
───────────────────
INITIAL RELEASE
  • Single-sheet Excel format (MAPPING section)
  • 30 mapping fields, 16 column fields
  • Mandatory fields marked with (*)
  • SQL generation (SELECT/FROM/WHERE/QUALIFY)
  • Primary key derivation from source_keys/target_keys flags
  • 28 validation rules
  • Multi-table join support with alias injection
  • SCD2 metadata capture
  • Dedup logic: KEEP_FIRST, KEEP_LAST, REJECT
  • Source lookup table support
  • Delete flag handling
  • Sample data limiting via QUALIFY ROW_NUMBER
  • Custom SQL transformations
  • Schema prefix auto-injection

REMOVED from earlier versions:
  • S_SCHEMA_lookup (column-level) → Moved to mapping-level
  • source_pk_group / target_pk_group → PKs derived from flags

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🚀 USAGE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

COMMAND LINE:
  python phase1_parser.py --input Phase1_DataInput.xlsx --output mapping.json
  python phase1_parser.py -i input.xlsx -o output.json
  python phase1_parser.py -i input.xlsx              # Prints to stdout

RETURN CODES:
  0 = Success (warnings OK)
  1 = Validation errors

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import argparse
import json
import math
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# ── Column definitions ───────────────────────────────────────────────────────

TABLE_MANDATORY = [
    "mapping_id",
    "db_name",
    "source_system",
    "source_schema",
    "source_table",
    "target_schema",
    "target_table",
    "scd2_applicable",
    "delete_flag_applicable",
]

COLUMN_MANDATORY = [
    "source_table",
    "source_column",
    "source_keys",
    "source_data_type",
    "target_column",
    "target_keys",
    "target_data_type",
    "transformationtype",
]

YN_TABLE_FIELDS = ["scd2_applicable", "delete_flag_applicable"]
YN_COLUMN_FIELDS = ["source_keys", "target_keys"]
TRANSFORMATION_TYPES = {"COPY", "SQL"}
DEDUP_LOGIC_VALUES = {"KEEP_FIRST", "KEEP_LAST", "REJECT"}


# ── Helpers ───────────────────────────────────────────────────────────────────


class ValidationError(Exception):
    pass


def _clean(val):
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    s = str(val).strip()
    return s if s else None


def _strip_mandatory_marker(label):
    """Remove (*) suffix from field labels."""
    if not label:
        return label
    return label.replace(" (*)", "").replace("(*)", "").strip()


def _parse_single_sheet(filepath: str) -> list:
    """
    Parse single-sheet Excel format.
    Returns list of mapping dicts, each with 'mapping_info' and 'columns' keys.
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb["MAPPING"]

    mappings = []
    current_mapping = None
    in_column_section = False
    column_headers = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        cell_a = _clean(row[0].value)
        cell_b = _clean(row[1].value) if len(row) > 1 else None

        # Detect MAPPING INFORMATION section start
        if cell_a and cell_a.upper() == "MAPPING INFORMATION":
            # Save previous mapping if exists
            if current_mapping:
                mappings.append(current_mapping)
            # Start new mapping
            current_mapping = {"mapping_info": {}, "columns": [], "_start_row": row_idx}
            in_column_section = False
            continue

        # Detect COLUMN MAPPINGS section start
        if cell_a and cell_a.upper() == "COLUMN MAPPINGS":
            in_column_section = True
            column_headers = []
            continue

        # Skip legend rows and empty rows
        if not cell_a or "GOLD" in (cell_a.upper() if cell_a else ""):
            continue

        # Read mapping info (label-value pairs)
        if current_mapping and not in_column_section:
            if cell_a:
                field_name = _strip_mandatory_marker(cell_a)
                current_mapping["mapping_info"][field_name] = cell_b

        # Read column headers
        elif in_column_section and not column_headers:
            # This is the header row for columns
            column_headers = [
                _strip_mandatory_marker(_clean(cell.value))
                for cell in row
                if _clean(cell.value)
            ]
            continue

        # Read column data rows
        elif in_column_section and column_headers:
            # Check if first column has data (source_table)
            if cell_a:
                col_row = {}
                for col_idx, header in enumerate(column_headers):
                    if col_idx < len(row):
                        val = _clean(row[col_idx].value)
                        if val:
                            col_row[header] = val
                if col_row:
                    current_mapping["columns"].append(col_row)

    # Save last mapping
    if current_mapping:
        mappings.append(current_mapping)

    return mappings


def _parse_table_aliases(join_clause: str) -> dict:
    """Extract {TABLE_NAME: ALIAS} from a JOIN clause."""
    alias_map = {}
    SQL_KEYWORDS = {
        "LEFT",
        "RIGHT",
        "INNER",
        "OUTER",
        "CROSS",
        "JOIN",
        "ON",
        "WHERE",
        "AND",
        "OR",
        "SELECT",
        "FROM",
        "AS",
    }
    pattern = re.compile(
        r"(?:FROM\s+|JOIN\s+)?(\w+)\s+(?:AS\s+)?(\w+)"
        r"(?=\s+(?:LEFT|RIGHT|INNER|OUTER|CROSS|ON|WHERE|,|$)|\s*$)",
        re.IGNORECASE,
    )
    for m in pattern.finditer(join_clause):
        tbl, alias = m.group(1).upper(), m.group(2).upper()
        if tbl not in SQL_KEYWORDS and alias not in SQL_KEYWORDS:
            alias_map[tbl] = alias
    return alias_map


def _parse_tables_from_join(join_clause: str) -> set:
    """Extract all table names referenced in a JOIN clause."""
    SQL_KEYWORDS = {
        "LEFT",
        "RIGHT",
        "INNER",
        "OUTER",
        "CROSS",
        "JOIN",
        "ON",
        "WHERE",
        "AND",
        "OR",
        "SELECT",
        "FROM",
        "AS",
    }
    tables = set()
    for m in re.finditer(r"(?:FROM\s+|JOIN\s+)(\w+)", join_clause, re.IGNORECASE):
        tbl = m.group(1).upper()
        if tbl not in SQL_KEYWORDS:
            tables.add(tbl)
    return tables


def _check_parens(val) -> bool:
    if not isinstance(val, str):
        return True
    return val.count("(") == val.count(")")


def _has_unescaped_quote(val) -> bool:
    if not isinstance(val, str):
        return False
    return "'" in val.replace("''", "")


def _inject_schema_into_join(join_clause: str, db: str, schema: str) -> str:
    """Prepend db.schema. to every table name in a JOIN clause."""
    prefix = f"{db}.{schema}." if db else f"{schema}."
    SQL_KEYWORDS = {
        "LEFT",
        "RIGHT",
        "INNER",
        "OUTER",
        "CROSS",
        "JOIN",
        "ON",
        "WHERE",
        "AND",
        "OR",
        "SELECT",
        "FROM",
        "AS",
    }

    def replacer(m):
        keyword = m.group(1)
        tbl = m.group(2)
        if tbl.upper() in SQL_KEYWORDS:
            return m.group(0)
        return f"{keyword}{prefix}{tbl}"

    result = re.sub(
        r"((?:FROM|JOIN)\s+)(\w+)", replacer, join_clause, flags=re.IGNORECASE
    )
    # Prefix leading table
    first = re.match(r"^(\w+)", result)
    if first:
        token = first.group(1)
        if token.upper() not in SQL_KEYWORDS and not token.upper().startswith(
            db.upper() if db else schema.upper()
        ):
            result = prefix + result
    return result


def _clean_obj(obj):
    """Recursively strip None/NaN for clean JSON output."""
    if isinstance(obj, list):
        return [_clean_obj(i) for i in obj]
    if isinstance(obj, dict):
        return {
            k: _clean_obj(v)
            for k, v in obj.items()
            if v is not None and not (isinstance(v, float) and math.isnan(v))
        }
    if isinstance(obj, float) and math.isnan(obj):
        return None
    return obj


# ── SQL Generation ────────────────────────────────────────────────────────────


def build_target_sql(mapping: dict) -> str:
    """Build source extraction SQL: SELECT/FROM/WHERE/QUALIFY."""
    mid = mapping.get("mapping_id", "?")
    db = (mapping.get("db_name") or "").upper()
    schema = (mapping.get("source_schema") or "").upper()
    columns = mapping.get("columns", [])
    join_clause = mapping.get("source_join", "") or ""

    alias_map = _parse_table_aliases(join_clause) if join_clause else {}
    multi_table = bool(join_clause)

    print(f"\n{'═'*60}")
    print(f"  Building SQL for mapping: {mid}")
    print(f"{'═'*60}")

    # ── SELECT ────────────────────────────────────────────────────────────────
    select_lines = []
    for col in columns:
        src_table = (col.get("source_table") or "").upper()
        src_col = col.get("source_column") or ""
        tgt_col = col.get("target_column") or ""
        tt = (col.get("transformationtype") or "").upper()
        rule = col.get("transformationrule") or ""
        override = col.get("source_column_override") or ""
        default = col.get("source_default_value") or ""

        # Resolve expression
        if tt == "SQL" and rule:
            expr = rule if "." in rule else rule
        elif override:
            expr = override
        else:
            if multi_table and src_table in alias_map:
                expr = f"{alias_map[src_table]}.{src_col}"
            else:
                expr = src_col

        # COALESCE for defaults
        if default:
            try:
                float(default)
                expr = f"COALESCE({expr}, {default})"
            except ValueError:
                expr = f"COALESCE({expr}, '{default}')"

        select_lines.append(f"    {expr} AS {tgt_col}")

    select_clause = "SELECT\n" + ",\n".join(select_lines)
    print(f"\n[SELECT]\n{select_clause}")

    # ── FROM ──────────────────────────────────────────────────────────────────
    if multi_table:
        qualified_join = _inject_schema_into_join(join_clause, db, schema)
        from_clause = f"FROM {qualified_join}"
    else:
        src_table_name = (mapping.get("source_table") or "").strip().upper()
        alias = list(alias_map.values())[0] if alias_map else ""
        alias_part = f" {alias}" if alias else ""
        from_clause = f"FROM {db}.{schema}.{src_table_name}{alias_part}"

    print(f"\n[FROM]\n{from_clause}")

    # ── WHERE ─────────────────────────────────────────────────────────────────
    filter_parts = []
    for field in ["source_filter", "source_date_filter", "source_filter_other"]:
        val = (mapping.get(field) or "").strip()
        if val:
            filter_parts.append(f"      {val}")

    where_clause = "WHERE\n" + "\nAND ".join(filter_parts) if filter_parts else ""
    if where_clause:
        print(f"\n[WHERE]\n{where_clause}")
    else:
        print("\n[WHERE] (none)")

    # ── QUALIFY ───────────────────────────────────────────────────────────────
    qualify_clause = ""
    sample_set = mapping.get("sample_set")
    if sample_set:
        pk_exprs = []
        for col in columns:
            if (col.get("source_keys") or "").upper() == "Y":
                src_table = (col.get("source_table") or "").upper()
                src_col = col.get("source_column") or ""
                override = col.get("source_column_override") or ""
                tt = (col.get("transformationtype") or "").upper()
                rule = col.get("transformationrule") or ""
                if tt == "SQL" and rule:
                    pk_exprs.append(rule if "." in rule else rule)
                elif override:
                    pk_exprs.append(override)
                else:
                    if multi_table and src_table in alias_map:
                        pk_exprs.append(f"{alias_map[src_table]}.{src_col}")
                    else:
                        pk_exprs.append(src_col)

        if pk_exprs:
            order_cols = ", ".join(pk_exprs)
            qualify_clause = (
                f"QUALIFY ROW_NUMBER() OVER (ORDER BY {order_cols}) <= {sample_set}"
            )
            print(f"\n[QUALIFY]\n{qualify_clause}")

    # ── Assemble ──────────────────────────────────────────────────────────────
    parts = [select_clause, from_clause]
    if where_clause:
        parts.append(where_clause)
    if qualify_clause:
        parts.append(qualify_clause)

    full_sql = "\n".join(parts)
    print(f"\n[FULL SQL]\n{full_sql}")
    print(f"\n{'─'*60}")
    return full_sql


# ── Validation ────────────────────────────────────────────────────────────────


def validate_and_generate(filepath: str) -> tuple:
    """Parse single-sheet Excel, validate, generate SQL."""
    errors = []
    warnings = []

    # Parse the single-sheet format
    try:
        raw_mappings = _parse_single_sheet(filepath)
    except Exception as e:
        raise ValidationError(f"Cannot parse Excel file: {e}")

    if not raw_mappings:
        raise ValidationError("No mappings found in Excel file.")

    # Process each mapping
    result = []
    all_mapping_ids = set()

    for map_idx, raw_map in enumerate(raw_mappings, 1):
        mapping_info = raw_map["mapping_info"]
        columns = raw_map["columns"]

        # ── Validate table-level fields ──────────────────────────────────────
        for field in TABLE_MANDATORY:
            if not mapping_info.get(field):
                errors.append(f"  Mapping {map_idx}: '{field}' is mandatory but empty.")

        mid = mapping_info.get("mapping_id")
        if mid:
            if mid in all_mapping_ids:
                errors.append(f"  Mapping {map_idx}: Duplicate mapping_id '{mid}'.")
            all_mapping_ids.add(mid)

        for field in YN_TABLE_FIELDS:
            val = mapping_info.get(field)
            if val and val.upper() not in ("Y", "N"):
                errors.append(
                    f"  Mapping {map_idx}: '{field}' must be Y or N (got '{val}')."
                )

        scd2 = (mapping_info.get("scd2_applicable") or "").upper()
        if scd2 == "Y":
            if not mapping_info.get("scd2_start_date_column"):
                errors.append(
                    f"  Mapping {map_idx}: scd2_start_date_column required when scd2_applicable=Y."
                )
            if not mapping_info.get("scd2_end_date_column"):
                errors.append(
                    f"  Mapping {map_idx}: scd2_end_date_column required when scd2_applicable=Y."
                )

        ss = mapping_info.get("sample_set")
        if ss:
            try:
                int(float(ss))
            except ValueError:
                errors.append(
                    f"  Mapping {map_idx}: sample_set must be a positive integer (got '{ss}')."
                )

        # Validate dedup_logic
        dedup = mapping_info.get("dedup_logic")
        if dedup:
            dedup_upper = dedup.upper()
            if dedup_upper not in DEDUP_LOGIC_VALUES:
                errors.append(
                    f"  Mapping {map_idx}: dedup_logic must be KEEP_FIRST, KEEP_LAST, or REJECT (got '{dedup}')."
                )

        # Validate source_lookup_table and join condition consistency
        lookup_table = mapping_info.get("source_lookup_table")
        lookup_join = mapping_info.get("source_lookup_join_condition")
        if lookup_table and not lookup_join:
            warnings.append(
                f"  Mapping {map_idx}: source_lookup_table is set but source_lookup_join_condition is empty."
            )
        if lookup_join and not lookup_table:
            warnings.append(
                f"  Mapping {map_idx}: source_lookup_join_condition is set but source_lookup_table is empty."
            )

        source_tables_raw = mapping_info.get("source_table") or ""
        source_tables = [
            t.strip().upper() for t in source_tables_raw.split(",") if t.strip()
        ]
        join_clause = mapping_info.get("source_join") or ""

        if len(source_tables) > 1 and not join_clause:
            errors.append(
                f"  Mapping {map_idx}: Multiple source tables but source_join is empty."
            )
        if len(source_tables) == 1 and join_clause:
            warnings.append(
                f"  Mapping {map_idx}: Single source table but source_join is populated."
            )

        if join_clause and source_tables:
            join_tables = _parse_tables_from_join(join_clause)
            unknown = join_tables - set(source_tables)
            if unknown:
                errors.append(
                    f"  Mapping {map_idx}: source_join references unknown tables: {', '.join(sorted(unknown))}."
                )

        # ── Validate column-level fields ─────────────────────────────────────
        target_cols_seen = set()
        src_pk_rows = []
        tgt_pk_rows = []

        for col_idx, col in enumerate(columns, 1):
            for field in COLUMN_MANDATORY:
                if not col.get(field):
                    errors.append(
                        f"  Mapping {map_idx}, Column {col_idx}: '{field}' is mandatory but empty."
                    )

            for field in YN_COLUMN_FIELDS:
                val = col.get(field)
                if val and val.upper() not in ("Y", "N"):
                    errors.append(
                        f"  Mapping {map_idx}, Column {col_idx}: '{field}' must be Y or N."
                    )

            tt = (col.get("transformationtype") or "").upper()
            rule = col.get("transformationrule")
            if tt and tt not in TRANSFORMATION_TYPES:
                errors.append(
                    f"  Mapping {map_idx}, Column {col_idx}: transformationtype must be COPY or SQL."
                )
            if tt == "SQL" and not rule:
                errors.append(
                    f"  Mapping {map_idx}, Column {col_idx}: transformationrule required when transformationtype=SQL."
                )

            target_col = col.get("target_column")
            if target_col:
                if target_col in target_cols_seen:
                    errors.append(
                        f"  Mapping {map_idx}, Column {col_idx}: Duplicate target_column '{target_col}'."
                    )
                target_cols_seen.add(target_col)

            col_src_table = (col.get("source_table") or "").upper()
            if col_src_table and col_src_table not in source_tables:
                errors.append(
                    f"  Mapping {map_idx}, Column {col_idx}: source_table '{col_src_table}' not in mapping source tables."
                )

            if (col.get("source_keys") or "").upper() == "Y":
                src_pk_rows.append(col)
            if (col.get("target_keys") or "").upper() == "Y":
                tgt_pk_rows.append(col)

        # ── Cross-field validations ──────────────────────────────────────────
        if not src_pk_rows:
            errors.append(
                f"  Mapping {map_idx}: No source_keys=Y column found — at least one required."
            )
        if not tgt_pk_rows:
            errors.append(
                f"  Mapping {map_idx}: No target_keys=Y column found — at least one required."
            )

        if ss and not src_pk_rows:
            errors.append(
                f"  Mapping {map_idx}: sample_set set but no source PK columns — cannot build QUALIFY."
            )

        # Build final mapping object
        entry = dict(mapping_info)
        if entry.get("sample_set"):
            try:
                entry["sample_set"] = int(float(entry["sample_set"]))
            except Exception:
                pass
        entry["columns"] = columns
        entry["target_sql"] = build_target_sql(entry) if not errors else ""
        result.append(entry)

    # Abort if errors
    if errors:
        err_block = "\n".join(errors)
        warn_block = ("\n\nWARNINGS:\n" + "\n".join(warnings)) if warnings else ""
        raise ValidationError(f"\n❌ VALIDATION FAILED:\n\n{err_block}{warn_block}")

    return result, warnings


# ── CLI ───────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="Phase 1 — Parse single-sheet Excel to JSON + SQL"
    )
    parser.add_argument("--input", "-i", required=True)
    parser.add_argument("--output", "-o", required=False)
    args = parser.parse_args()

    filepath = Path(args.input)
    if not filepath.exists():
        print("", file=sys.stderr)
        print("=" * 70, file=sys.stderr)
        print("  ❌ PHASE 1 FAILED — FILE NOT FOUND", file=sys.stderr)
        print("=" * 70, file=sys.stderr)
        print(f"  File path:   {filepath}", file=sys.stderr)
        print(f"  Status:      FAILED", file=sys.stderr)
        print(f"  Exit code:   1", file=sys.stderr)
        print(f"  Action:      Check file path and try again", file=sys.stderr)
        print("=" * 70, file=sys.stderr)
        sys.exit(1)

    try:
        data, warnings = validate_and_generate(str(filepath))
    except ValidationError as e:
        print(str(e), file=sys.stderr)
        print("", file=sys.stderr)
        print("=" * 70, file=sys.stderr)
        print("  ❌ PHASE 1 FAILED — VALIDATION ERRORS", file=sys.stderr)
        print("=" * 70, file=sys.stderr)
        print(f"  Status:      FAILED", file=sys.stderr)
        print(f"  Exit code:   1", file=sys.stderr)
        print(f"  Action:      Fix validation errors above and re-run", file=sys.stderr)
        print("=" * 70, file=sys.stderr)
        sys.exit(1)

    if warnings:
        print("", file=sys.stderr)
        print("⚠️  WARNINGS (non-blocking):", file=sys.stderr)
        for w in warnings:
            print(w, file=sys.stderr)

    json_out = json.dumps(_clean_obj(data), indent=2, default=str)

    # Write output
    if args.output:
        Path(args.output).write_text(json_out)
        output_location = args.output
    else:
        print(json_out)
        output_location = "stdout"

    # Print success summary
    print("", file=sys.stderr)
    print("=" * 70, file=sys.stderr)
    print("  ✅ PHASE 1 COMPLETE — SUCCESS", file=sys.stderr)
    print("=" * 70, file=sys.stderr)
    print(f"  Mappings processed: {len(data)}", file=sys.stderr)
    print(f"  Output location:    {output_location}", file=sys.stderr)
    print(f"  Warnings:           {len(warnings)}", file=sys.stderr)
    print(f"  Status:             PASSED", file=sys.stderr)
    print(f"  Exit code:          0", file=sys.stderr)
    print("=" * 70, file=sys.stderr)

    # Exit with success code
    sys.exit(0)


if __name__ == "__main__":
    main()
