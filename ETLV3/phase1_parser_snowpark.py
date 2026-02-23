"""
================================================================================
              PHASE 1 - DATA INPUT PARSER (Snowpark Worksheet)
                        (Single-Sheet Format)
================================================================================

VERSION:        1.1.3
RELEASE DATE:   2026-02-19
AUTHOR:         QA
LANGUAGE:       Python 3.9+
DEPENDENCIES:   pandas, openpyxl (install via Packages menu in Snowsight)

HOW TO RUN IN SNOWFLAKE:
  1. Upload Phase1_DataInput.xlsx to Snowflake stage
     SQL> PUT file:///path/to/Phase1_DataInput.xlsx @stgintegration AUTO_COMPRESS=FALSE;

  2. Open a Snowpark Python worksheet in Snowsight

  3. Add packages (click "Packages" dropdown top-right):
     - pandas
     - openpyxl

  4. Paste this entire script into the worksheet

  5. Edit CONFIGURATION section below

  6. Click "Run" button

OUTPUT:
  - Writes to Snowflake table: DATABASE.SCHEMA.TABLE_NAME
  - Table columns: mapping_id, run_timestamp, mapping_json, source_sql, target_sql
  - Console shows step-by-step progress and SQL generation

CHANGES in v1.1.3:
  - Auto-casting when source/target data types differ
  - source_sql: Extracts FROM source tables (renamed from target_sql)
  - target_sql: Queries FROM target tables (NEW)
  - Empty source_column allowed for system/hardcoded columns
  - Removed graphic icons for Snowflake compatibility

================================================================================
"""

import json
import math
import re
import tempfile
import os
import gzip
import shutil

import pandas as pd
from snowflake.snowpark import Session

# ==============================================================================
# [CONFIG]  CONFIGURATION - EDIT THESE VALUES
# ==============================================================================

# INPUT: Snowflake stage path to Excel file
STAGE_INPUT_PATH = "@stgintegration/Phase1_DataInput.xlsx"

# OUTPUT: Snowflake table to write results
OUTPUT_DATABASE = None  # None = use current database
OUTPUT_SCHEMA = "PUBLIC"  # Schema name
OUTPUT_TABLE = "PHASE1_MAPPING"  # Table name

# BEHAVIOR
WRITE_TO_TABLE = True  # Set False to skip table write (dry run)

# ==============================================================================


# Column definitions ------------------------------------------------------------

TABLE_MANDATORY = [
    "mapping_id",
    "db_name",
    "source_system",
    "source_schema",
    "source_table",
    "target_schema",
    "target_table",
    "scd2_applicable",
    "delete_flag_applicable",
]

COLUMN_MANDATORY = [
    "target_column",
    "target_keys",
    "target_data_type",
    "transformationtype",
]

# Optional fields that can be empty:
# - source_table (for system/hardcoded columns)
# - source_column (for system/hardcoded columns)
# - source_keys (for non-PK columns)
# - source_data_type (for system columns)

# Note: source_column, source_table can be empty when transformationtype=SQL
# This allows for system columns like CURRENT_TIMESTAMP(), hardcoded values, etc.

YN_TABLE_FIELDS = ["scd2_applicable", "delete_flag_applicable"]
YN_COLUMN_FIELDS = ["source_keys", "target_keys"]
TRANSFORMATION_TYPES = {"COPY", "SQL"}
DEDUP_LOGIC_VALUES = {"KEEP_FIRST", "KEEP_LAST", "REJECT"}


# Helpers -----------------------------------------------------------------------


class ValidationError(Exception):
    pass


def _clean(val):
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    s = str(val).strip()
    return s if s else None


def _strip_mandatory_marker(label):
    if not label:
        return label
    return label.replace(" (*)", "").replace("(*)", "").strip()


def _parse_single_sheet(filepath: str) -> list:
    """Parse single-sheet Excel format - with comprehensive debugging."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, data_only=True)
    ws = wb["MAPPING"]

    mappings = []
    current_mapping = None
    in_column_section = False
    column_headers = []

    print("" + "=" * 80)
    print("EXCEL PARSE DEBUG - Starting")
    print("=" * 80)

    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        cell_a = _clean(row[0].value)
        cell_b = _clean(row[1].value) if len(row) > 1 else None

        # Debug output for rows 34-50
        if 34 <= row_idx <= 50:
            all_values = [_clean(cell.value) for cell in row[:15]]
            non_empty = [v for v in all_values if v]
            print(
                f"Row {row_idx}: First cell='{cell_a}', Non-empty cells={len(non_empty)}"
            )

        # Check for MAPPING INFORMATION section
        if cell_a and cell_a.upper() == "MAPPING INFORMATION":
            if row_idx >= 34:
                print(f"  ACTION: Starting new mapping section")
            if current_mapping:
                mappings.append(current_mapping)
            current_mapping = {"mapping_info": {}, "columns": [], "_start_row": row_idx}
            in_column_section = False
            continue

        # Check for COLUMN MAPPINGS section
        if cell_a and "COLUMN" in cell_a.upper() and "MAPPING" in cell_a.upper():
            if row_idx >= 34:
                print(f"  ACTION: Entering column section")
            in_column_section = True
            column_headers = []
            continue

        # Skip empty rows ONLY if we're not in column section with headers
        # In column section, rows can have empty first cell (for system columns)
        if not cell_a:
            if not (in_column_section and column_headers):
                # Not in column data section, safe to skip
                if 37 <= row_idx <= 50:
                    print(
                        f"  ACTION: Skipped (empty first cell, not in column section)"
                    )
                continue
            # else: We're in column section with headers, DON'T skip - process the row

        # Skip legend and color explanation rows (only check if cell_a has value)
        if cell_a:
            cell_a_upper = cell_a.upper()
            skip_keywords = ["GOLD", "BLUE", "MANDATORY", "OPTIONAL", "LEGEND"]
            if any(keyword in cell_a_upper for keyword in skip_keywords):
                if row_idx >= 34:
                    matched = [k for k in skip_keywords if k in cell_a_upper]
                    print(f"  ACTION: Skipped (legend row, matched: {matched})")
                continue

        # Process mapping info fields
        if current_mapping and not in_column_section:
            if cell_a:
                field_name = _strip_mandatory_marker(cell_a)
                current_mapping["mapping_info"][field_name] = cell_b

        # Get column headers
        elif in_column_section and not column_headers:
            headers_temp = [
                _strip_mandatory_marker(_clean(cell.value))
                for cell in row
                if _clean(cell.value)
            ]

            if row_idx >= 34:
                print(f"  CHECKING HEADERS: {headers_temp[:5]}...")

            # Verify we got actual column headers
            expected = ["source_table", "target_column", "transformationtype"]
            has_expected = any(h in expected for h in headers_temp)

            if headers_temp and has_expected:
                column_headers = headers_temp
                print(f"  ACTION: CONFIRMED {len(column_headers)} headers")
                print(f"          Headers: {column_headers}")
                continue
            else:
                if row_idx >= 34:
                    print(f"  ACTION: Not headers (missing expected columns)")
                continue

        # Process column data rows
        elif in_column_section and column_headers:
            col_row = {}
            for col_idx, header in enumerate(column_headers):
                if col_idx < len(row):
                    val = _clean(row[col_idx].value)
                    if val:
                        col_row[header] = val

            if row_idx >= 37:
                print(f"  DATA ROW: {len(col_row)} fields populated")
                print(
                    f"           target_column='{col_row.get('target_column', '(missing)')}'"
                )
                print(f"           Fields: {list(col_row.keys())[:5]}...")

            # Pick up row if it has target_column
            if col_row and col_row.get("target_column"):
                if row_idx >= 37:
                    print(
                        f"  ACTION: ✅ PICKED UP (target='{col_row.get('target_column')}')"
                    )
                current_mapping["columns"].append(col_row)
            else:
                if row_idx >= 37:
                    if not col_row:
                        print(f"  ACTION: ❌ SKIPPED (col_row is empty)")
                    else:
                        print(f"  ACTION: ❌ SKIPPED (no target_column)")

    if current_mapping:
        mappings.append(current_mapping)

    print("\n" + "=" * 80)
    print("EXCEL PARSE DEBUG - Summary")
    print("=" * 80)
    for m_idx, m in enumerate(mappings, 1):
        print(f"Mapping {m_idx}: {len(m['columns'])} columns parsed")
        for c_idx, c in enumerate(m["columns"], 1):
            src = c.get("source_column", "(empty)")
            tgt = c.get("target_column", "(empty)")
            tt = c.get("transformationtype", "?")
            print(f"  {c_idx}. {src} -> {tgt} [{tt}]")
    print("=" * 80 + "\n")

    return mappings


def _parse_table_aliases(join_clause: str) -> dict:
    alias_map = {}
    SQL_KEYWORDS = {
        "LEFT",
        "RIGHT",
        "INNER",
        "OUTER",
        "CROSS",
        "JOIN",
        "ON",
        "WHERE",
        "AND",
        "OR",
        "SELECT",
        "FROM",
        "AS",
    }
    pattern = re.compile(
        r"(?:FROM\s+|JOIN\s+)?(\w+)\s+(?:AS\s+)?(\w+)"
        r"(?=\s+(?:LEFT|RIGHT|INNER|OUTER|CROSS|ON|WHERE|,|$)|\s*$)",
        re.IGNORECASE,
    )
    for m in pattern.finditer(join_clause):
        tbl, alias = m.group(1).upper(), m.group(2).upper()
        if tbl not in SQL_KEYWORDS and alias not in SQL_KEYWORDS:
            alias_map[tbl] = alias
    return alias_map


def _parse_tables_from_join(join_clause: str) -> set:
    SQL_KEYWORDS = {
        "LEFT",
        "RIGHT",
        "INNER",
        "OUTER",
        "CROSS",
        "JOIN",
        "ON",
        "WHERE",
        "AND",
        "OR",
        "SELECT",
        "FROM",
        "AS",
    }
    tables = set()
    for m in re.finditer(r"(?:FROM\s+|JOIN\s+)(\w+)", join_clause, re.IGNORECASE):
        tbl = m.group(1).upper()
        if tbl not in SQL_KEYWORDS:
            tables.add(tbl)
    return tables


def _check_parens(val) -> bool:
    if not isinstance(val, str):
        return True
    return val.count("(") == val.count(")")


def _has_unescaped_quote(val) -> bool:
    if not isinstance(val, str):
        return False
    return "'" in val.replace("''", "")


def _inject_schema_into_join(join_clause: str, db: str, schema: str) -> str:
    prefix = f"{db}.{schema}." if db else f"{schema}."
    SQL_KEYWORDS = {
        "LEFT",
        "RIGHT",
        "INNER",
        "OUTER",
        "CROSS",
        "JOIN",
        "ON",
        "WHERE",
        "AND",
        "OR",
        "SELECT",
        "FROM",
        "AS",
    }

    def replacer(m):
        keyword = m.group(1)
        tbl = m.group(2)
        if tbl.upper() in SQL_KEYWORDS:
            return m.group(0)
        return f"{keyword}{prefix}{tbl}"

    result = re.sub(
        r"((?:FROM|JOIN)\s+)(\w+)", replacer, join_clause, flags=re.IGNORECASE
    )
    first = re.match(r"^(\w+)", result)
    if first:
        token = first.group(1)
        if token.upper() not in SQL_KEYWORDS and not token.upper().startswith(
            db.upper() if db else schema.upper()
        ):
            result = prefix + result
    return result


def _clean_obj(obj):
    if isinstance(obj, list):
        return [_clean_obj(i) for i in obj]
    if isinstance(obj, dict):
        return {
            k: _clean_obj(v)
            for k, v in obj.items()
            if v is not None and not (isinstance(v, float) and math.isnan(v))
        }
    if isinstance(obj, float) and math.isnan(obj):
        return None
    return obj


def _fix_string_quotes(value: str) -> str:
    """
    Fix string literals that may have been mangled by Excel.
    Handles:
    - SAP' -> 'SAP' (Excel strips leading quote)
    - "SAP" -> 'SAP' (convert double quotes)
    - 'SAP' -> 'SAP' (already correct)
    - SAP -> 'SAP' (add quotes for bare word)

    Does NOT modify:
    - Column references: T1.COLUMN_NAME
    - Functions: CURRENT_TIMESTAMP()
    - SQL keywords: NULL, TRUE, FALSE
    - Numbers: 123, 45.67
    - Complex expressions with operators
    """
    if not value:
        return value

    value = value.strip()

    # Don't modify if it contains SQL operators or functions
    sql_indicators = [
        "(",
        ")",
        ".",
        "CASE",
        "WHEN",
        "THEN",
        "END",
        "CAST",
        "COALESCE",
        "||",
        "+",
        "-",
        "*",
        "/",
        "SELECT",
        "FROM",
        "WHERE",
    ]
    if any(ind in value.upper() for ind in sql_indicators):
        return value

    # Don't modify SQL keywords
    if value.upper() in [
        "NULL",
        "TRUE",
        "FALSE",
        "CURRENT_TIMESTAMP",
        "CURRENT_DATE",
        "CURRENT_TIME",
        "CURRENT_USER",
    ]:
        return value

    # Don't modify numbers
    try:
        float(value.replace(",", ""))
        return value
    except ValueError:
        pass

    # Case 1: Double quotes -> convert to single quotes
    if value.startswith('"') and value.endswith('"'):
        inner = value[1:-1]
        # Escape any single quotes in the inner string
        inner = inner.replace("'", "''")
        return f"'{inner}'"

    # Case 2: Missing leading quote (Excel ate it) - ends with quote but doesn't start
    if value.endswith("'") and not value.startswith("'"):
        # Add the missing leading quote
        return f"'{value}"

    # Case 3: Already has both quotes
    if value.startswith("'") and value.endswith("'"):
        return value

    # Case 4: Bare word that looks like a string literal (no spaces, not a column ref)
    # Only add quotes if it's a simple word/text (no dots, not uppercase SQL keywords)
    if " " not in value and "." not in value and not value.isupper():
        return f"'{value}'"

    # Default: return as-is (might be column reference or complex expression)
    return value


# SQL Generation ----------------------------------------------------------------


def build_source_sql(mapping: dict) -> str:
    """Build SQL to extract FROM source tables."""
    mid = mapping.get("mapping_id", "?")
    db = (mapping.get("db_name") or "").upper()
    schema = (mapping.get("source_schema") or "").upper()
    columns = mapping.get("columns", [])
    join_clause = mapping.get("source_join", "") or ""

    alias_map = _parse_table_aliases(join_clause) if join_clause else {}
    multi_table = bool(join_clause)

    print(f"\n{'='*60}")
    print(f"  Building SOURCE SQL for mapping: {mid}")
    print(f"{'='*60}")

    select_lines = []
    for col in columns:
        src_table = (col.get("source_table") or "").upper()
        src_col = col.get("source_column") or ""
        tgt_col = col.get("target_column") or ""
        tt = (col.get("transformationtype") or "").upper()
        rule = col.get("transformationrule") or ""
        override = col.get("source_column_override") or ""
        default = col.get("source_default_value") or ""

        if tt == "SQL" and rule:
            # SQL transformation - fix quotes if needed
            expr = _fix_string_quotes(rule)
        elif override:
            # Use column override
            expr = override
        else:
            # COPY mode
            if not src_col:
                # Empty source_column - should be SQL type but handle gracefully
                print(
                    f"  [WARNING] Column {tgt_col}: source_column is empty, using NULL"
                )
                expr = "NULL"
            elif not src_table:
                # Empty source_table - use column without prefix
                expr = src_col
            elif multi_table and src_table in alias_map:
                expr = f"{alias_map[src_table]}.{src_col}"
            else:
                expr = src_col

        # Apply default value
        if default and expr != "NULL":
            try:
                float(default)
                expr = f"COALESCE({expr}, {default})"
            except ValueError:
                expr = f"COALESCE({expr}, '{default}')"

        # Auto-cast if source and target data types differ
        src_dtype = (col.get("source_data_type") or "").upper()
        tgt_dtype = (col.get("target_data_type") or "").upper()

        if src_dtype and tgt_dtype and src_dtype != tgt_dtype:
            # Only apply casting for COPY or simple expressions (not SQL transformations with rules)
            if tt != "SQL" or not rule:
                if "DATE" in tgt_dtype and "DATE" not in src_dtype:
                    expr = f"TRY_TO_DATE({expr})"
                elif "TIMESTAMP" in tgt_dtype and "TIMESTAMP" not in src_dtype:
                    expr = f"TRY_TO_TIMESTAMP({expr})"
                elif (
                    "NUMBER" in tgt_dtype
                    or "NUMERIC" in tgt_dtype
                    or "DECIMAL" in tgt_dtype
                ):
                    expr = f"TRY_TO_NUMBER({expr})"
                elif (
                    "INTEGER" in tgt_dtype
                    or "INT" in tgt_dtype
                    or "BIGINT" in tgt_dtype
                ):
                    expr = f"TRY_TO_NUMBER({expr})"
                elif (
                    "VARCHAR" in tgt_dtype
                    or "STRING" in tgt_dtype
                    or "TEXT" in tgt_dtype
                ):
                    expr = f"TO_VARCHAR({expr})"
                elif "BOOLEAN" in tgt_dtype or "BOOL" in tgt_dtype:
                    expr = f"TRY_TO_BOOLEAN({expr})"

        select_lines.append(f"    {expr} AS {tgt_col}")

    select_clause = "SELECT\n" + ",\n".join(select_lines)
    print(f"\n[SELECT]\n{select_clause}")

    # FROM
    if multi_table:
        qualified_join = _inject_schema_into_join(join_clause, db, schema)
        from_clause = f"FROM {qualified_join}"
    else:
        src_table_name = (mapping.get("source_table") or "").strip().upper()
        alias = list(alias_map.values())[0] if alias_map else ""
        alias_part = f" {alias}" if alias else ""
        from_clause = f"FROM {db}.{schema}.{src_table_name}{alias_part}"

    print(f"\n[FROM]\n{from_clause}")

    # WHERE
    filter_parts = []
    for field in ["source_filter", "source_date_filter", "source_filter_other"]:
        val = (mapping.get(field) or "").strip()
        if val:
            filter_parts.append(f"      {val}")

    where_clause = "WHERE\n" + "\nAND ".join(filter_parts) if filter_parts else ""
    if where_clause:
        print(f"\n[WHERE]\n{where_clause}")
    else:
        print("\n[WHERE] (none)")

    # QUALIFY
    qualify_clause = ""
    sample_set = mapping.get("sample_set")
    if sample_set:
        pk_exprs = []
        for col in columns:
            if (col.get("source_keys") or "").upper() == "Y":
                src_table = (col.get("source_table") or "").upper()
                src_col = col.get("source_column") or ""
                override = col.get("source_column_override") or ""
                tt = (col.get("transformationtype") or "").upper()
                rule = col.get("transformationrule") or ""
                if tt == "SQL" and rule:
                    pk_exprs.append(rule)
                elif override:
                    pk_exprs.append(override)
                elif src_col:
                    if multi_table and src_table in alias_map:
                        pk_exprs.append(f"{alias_map[src_table]}.{src_col}")
                    else:
                        pk_exprs.append(src_col)

        if pk_exprs:
            order_cols = ", ".join(pk_exprs)
            qualify_clause = (
                f"QUALIFY ROW_NUMBER() OVER (ORDER BY {order_cols}) <= {sample_set}"
            )
            print(f"\n[QUALIFY]\n{qualify_clause}")

    parts = [select_clause, from_clause]
    if where_clause:
        parts.append(where_clause)
    if qualify_clause:
        parts.append(qualify_clause)

    full_sql = "\n".join(parts)
    print(f"\n[FULL SOURCE SQL]\n{full_sql}")
    print(f"\n{'-'*60}")
    return full_sql


def build_target_sql(mapping: dict) -> str:
    """Build SQL to query FROM target table."""
    mid = mapping.get("mapping_id", "?")
    db = (mapping.get("db_name") or "").upper()
    schema = (mapping.get("target_schema") or "").upper()
    table = (mapping.get("target_table") or "").upper()
    columns = mapping.get("columns", [])
    join_clause = mapping.get("target_join", "") or ""

    print(f"\n{'='*60}")
    print(f"  Building TARGET SQL for mapping: {mid}")
    print(f"{'='*60}")

    # SELECT - use target column names
    select_lines = []
    for col in columns:
        tgt_col = col.get("target_column") or ""
        if tgt_col:
            select_lines.append(f"    {tgt_col}")

    if not select_lines:
        select_lines.append("    *")

    select_clause = "SELECT\n" + ",\n".join(select_lines)
    print(f"\n[SELECT]\n{select_clause}")

    # FROM
    if join_clause:
        from_clause = f"FROM {db}.{schema}.{join_clause}"
    else:
        from_clause = f"FROM {db}.{schema}.{table}"

    print(f"\n[FROM]\n{from_clause}")

    # WHERE
    filter_parts = []
    for field in ["target_filter", "target_date_filter", "target_filter_other"]:
        val = (mapping.get(field) or "").strip()
        if val:
            filter_parts.append(f"      {val}")

    where_clause = "WHERE\n" + "\nAND ".join(filter_parts) if filter_parts else ""
    if where_clause:
        print(f"\n[WHERE]\n{where_clause}")
    else:
        print("\n[WHERE] (none)")

    parts = [select_clause, from_clause]
    if where_clause:
        parts.append(where_clause)

    full_sql = "\n".join(parts)
    print(f"\n[FULL TARGET SQL]\n{full_sql}")
    print(f"\n{'-'*60}")
    return full_sql


# Validation --------------------------------------------------------------------


def validate_and_generate(filepath: str) -> tuple:
    errors = []
    warnings = []

    print("" + "=" * 60)
    print("  PHASE 1: VALIDATION & SQL GENERATION")
    print("=" * 60)

    try:
        raw_mappings = _parse_single_sheet(filepath)
    except Exception as e:
        raise ValidationError(f"Cannot parse Excel file: {e}")

    if not raw_mappings:
        raise ValidationError("No mappings found in Excel file.")

    print(f"[PARSE] Found {len(raw_mappings)} mapping(s) in Excel")

    result = []
    all_mapping_ids = set()

    for map_idx, raw_map in enumerate(raw_mappings, 1):
        mapping_info = raw_map["mapping_info"]
        columns = raw_map["columns"]

        print(f"[PARSE] Mapping {map_idx}: Found {len(columns)} column(s)")

        # Table-level validations
        print(f"[VALIDATE] Mapping {map_idx} ({mid}): Table-level checks...")
        for field in TABLE_MANDATORY:
            if not mapping_info.get(field):
                errors.append(f"  Mapping {map_idx}: '{field}' is mandatory but empty.")

        # mid already extracted earlier, just use it
        if mid:
            if mid in all_mapping_ids:
                errors.append(f"  Mapping {map_idx}: Duplicate mapping_id '{mid}'.")
            all_mapping_ids.add(mid)

        for field in YN_TABLE_FIELDS:
            val = mapping_info.get(field)
            if val and val.upper() not in ("Y", "N"):
                errors.append(
                    f"  Mapping {map_idx}: '{field}' must be Y or N (got '{val}')."
                )

        scd2 = (mapping_info.get("scd2_applicable") or "").upper()
        if scd2 == "Y":
            if not mapping_info.get("scd2_start_date_column"):
                errors.append(
                    f"  Mapping {map_idx}: scd2_start_date_column required when scd2_applicable=Y."
                )
            if not mapping_info.get("scd2_end_date_column"):
                errors.append(
                    f"  Mapping {map_idx}: scd2_end_date_column required when scd2_applicable=Y."
                )

        ss = mapping_info.get("sample_set")
        if ss:
            try:
                int(float(ss))
            except ValueError:
                errors.append(
                    f"  Mapping {map_idx}: sample_set must be a positive integer (got '{ss}')."
                )

        dedup = mapping_info.get("dedup_logic")
        if dedup:
            dedup_upper = dedup.upper()
            if dedup_upper not in DEDUP_LOGIC_VALUES:
                errors.append(
                    f"  Mapping {map_idx}: dedup_logic must be KEEP_FIRST, KEEP_LAST, or REJECT (got '{dedup}')."
                )

        lookup_table = mapping_info.get("source_lookup_table")
        lookup_join = mapping_info.get("source_lookup_join_condition")
        if lookup_table and not lookup_join:
            warnings.append(
                f"  Mapping {map_idx}: source_lookup_table is set but source_lookup_join_condition is empty."
            )
        if lookup_join and not lookup_table:
            warnings.append(
                f"  Mapping {map_idx}: source_lookup_join_condition is set but source_lookup_table is empty."
            )

        source_tables_raw = mapping_info.get("source_table") or ""
        source_tables = [
            t.strip().upper() for t in source_tables_raw.split(",") if t.strip()
        ]
        join_clause = mapping_info.get("source_join") or ""

        if len(source_tables) > 1 and not join_clause:
            errors.append(
                f"  Mapping {map_idx}: Multiple source tables but source_join is empty."
            )
        if len(source_tables) == 1 and join_clause:
            warnings.append(
                f"  Mapping {map_idx}: Single source table but source_join is populated."
            )

        if join_clause and source_tables:
            join_tables = _parse_tables_from_join(join_clause)
            unknown = join_tables - set(source_tables)
            if unknown:
                errors.append(
                    f"  Mapping {map_idx}: source_join references unknown tables: {', '.join(sorted(unknown))}."
                )

        # Column-level validations
        print(f"[VALIDATE] Mapping {map_idx}: Column-level checks...")
        target_cols_seen = set()
        src_pk_rows = []
        tgt_pk_rows = []

        for col_idx, col in enumerate(columns, 1):
            # Show progress for each column
            tgt_col = col.get("target_column", "?")
            src_col = col.get("source_column", "(empty)")
            tt = (col.get("transformationtype") or "").upper()
            print(f"  Column {col_idx}: {src_col} -> {tgt_col} [{tt}]")
            # Check if this is a system/hardcoded column
            tt = (col.get("transformationtype") or "").upper()
            is_system_column = (not col.get("source_column")) and (tt == "SQL")

            for field in COLUMN_MANDATORY:
                # Check only the truly mandatory fields
                if not col.get(field):
                    errors.append(
                        f"  Mapping {map_idx}, Column {col_idx}: '{field}' is mandatory but empty."
                    )

            # Additional conditional validations
            # If transformationtype=COPY, source_column should be present
            if tt == "COPY" and not col.get("source_column"):
                warnings.append(
                    f"  Mapping {map_idx}, Column {col_idx}: transformationtype=COPY but source_column is empty - consider using SQL."
                )

            for field in YN_COLUMN_FIELDS:
                val = col.get(field)
                if val and val.upper() not in ("Y", "N"):
                    errors.append(
                        f"  Mapping {map_idx}, Column {col_idx}: '{field}' must be Y or N."
                    )

            if tt and tt not in TRANSFORMATION_TYPES:
                errors.append(
                    f"  Mapping {map_idx}, Column {col_idx}: transformationtype must be COPY or SQL."
                )
            if tt == "SQL" and not col.get("transformationrule"):
                errors.append(
                    f"  Mapping {map_idx}, Column {col_idx}: transformationrule required when transformationtype=SQL."
                )

            target_col = col.get("target_column")
            if target_col:
                if target_col in target_cols_seen:
                    errors.append(
                        f"  Mapping {map_idx}, Column {col_idx}: Duplicate target_column '{target_col}'."
                    )
                target_cols_seen.add(target_col)

            col_src_table = (col.get("source_table") or "").upper()
            if col_src_table and col_src_table not in source_tables:
                errors.append(
                    f"  Mapping {map_idx}, Column {col_idx}: source_table '{col_src_table}' not in mapping source tables."
                )

            if (col.get("source_keys") or "").upper() == "Y":
                src_pk_rows.append(col)
            if (col.get("target_keys") or "").upper() == "Y":
                tgt_pk_rows.append(col)

        # Cross-field validations
        if not src_pk_rows:
            errors.append(
                f"  Mapping {map_idx}: No source_keys=Y column found - at least one required."
            )
        if not tgt_pk_rows:
            errors.append(
                f"  Mapping {map_idx}: No target_keys=Y column found - at least one required."
            )

        if ss and not src_pk_rows:
            errors.append(
                f"  Mapping {map_idx}: sample_set set but no source PK columns - cannot build QUALIFY."
            )

        # Build final mapping object
        entry = dict(mapping_info)
        if entry.get("sample_set"):
            try:
                entry["sample_set"] = int(float(entry["sample_set"]))
            except Exception:
                pass
        entry["columns"] = columns

        print(f"[VALIDATE] Mapping {map_idx}: Validation complete")
        print(f"           Columns: {len(columns)}")
        print(f"           Source PKs: {len(src_pk_rows)}")
        print(f"           Target PKs: {len(tgt_pk_rows)}")
        print(
            f"           Errors: {len([e for e in errors if f'Mapping {map_idx}' in e])}"
        )

        entry["source_sql"] = build_source_sql(entry) if not errors else ""
        entry["target_sql"] = build_target_sql(entry) if not errors else ""
        result.append(entry)

    if errors:
        err_block = "\n".join(errors)
        warn_block = ("\n\nWARNINGS:\n" + "\n".join(warnings)) if warnings else ""
        raise ValidationError(
            f"\n[ERROR] VALIDATION FAILED:\n\n{err_block}{warn_block}"
        )

    # Validation summary
    print("\n" + "=" * 60)
    print("  VALIDATION SUMMARY")
    print("=" * 60)
    print(f"  Total Mappings: {len(result)}")
    print(f"  Total Columns:  {sum(len(m['columns']) for m in result)}")
    print(f"  Errors:         0")
    print(f"  Warnings:       {len(warnings)}")
    print("=" * 60)

    return result, warnings


# SNOWPARK MAIN -----------------------------------------------------------------


def main(session: Session):
    """
    Snowpark worksheet entry point.
    Automatically called by Snowsight when you click Run.
    """
    log = []

    # Step 1: Download Excel from stage
    log.append(f"[STEP 1] Downloading {STAGE_INPUT_PATH} from stage...")
    tmp_dir = tempfile.mkdtemp()

    try:
        session.file.get(STAGE_INPUT_PATH, tmp_dir)
    except Exception as e:
        raise Exception(f"Cannot download from stage {STAGE_INPUT_PATH}: {e}")

    downloaded = [
        f for f in os.listdir(tmp_dir) if f.endswith((".xlsx", ".xls", ".gz"))
    ]
    if not downloaded:
        raise Exception(f"File not found after GET from stage: {STAGE_INPUT_PATH}")

    local_path = os.path.join(tmp_dir, downloaded[0])
    if local_path.endswith(".gz"):
        unzipped = local_path.replace(".gz", "")
        with gzip.open(local_path, "rb") as f_in, open(unzipped, "wb") as f_out:
            shutil.copyfileobj(f_in, f_out)
        local_path = unzipped

    log.append(f"[STEP 1] Downloaded: {local_path}  [OK]")

    # Step 2: Validate and generate SQL
    log.append("[STEP 2] Validating, parsing, and generating SQL...")
    try:
        data, warnings = validate_and_generate(local_path)
    except ValidationError as e:
        raise Exception(str(e))

    log.append(f"[STEP 2] Parsed {len(data)} mapping(s) successfully  [OK]")
    for w in warnings:
        log.append(f"  [WARNING]  {w}")

    # Step 3: Serialize to JSON
    json_str = json.dumps(_clean_obj(data), indent=2, default=str)
    log.append(f"[STEP 3] JSON serialized - {len(json_str):,} characters  [OK]")

    # Step 4: Write to Snowflake table
    if WRITE_TO_TABLE:
        from datetime import datetime
        from snowflake.snowpark.functions import parse_json, lit, current_timestamp
        from snowflake.snowpark.types import (
            StructType,
            StructField,
            StringType,
            TimestampType,
            VariantType,
        )

        db_prefix = f"{OUTPUT_DATABASE}." if OUTPUT_DATABASE else ""
        full_table = f"{db_prefix}{OUTPUT_SCHEMA}.{OUTPUT_TABLE}"
        log.append(f"[STEP 4] Writing to {full_table}...")

        # Drop and recreate table to ensure schema matches
        session.sql(f"""
            CREATE OR REPLACE TABLE {full_table} (
                mapping_id       VARCHAR,
                run_timestamp    TIMESTAMP_NTZ,
                mapping_json     VARIANT,
                source_sql       VARCHAR,
                target_sql       VARCHAR
            )
        """).collect()

        # Prepare data as list of tuples
        rows = []
        for mapping in data:
            rows.append(
                (
                    mapping.get("mapping_id", "UNKNOWN"),
                    json.dumps(_clean_obj(mapping), default=str),
                    mapping.get("source_sql") or "",
                    mapping.get("target_sql") or "",
                )
            )

        # Define schema for staging DataFrame
        schema = StructType(
            [
                StructField("mapping_id", StringType()),
                StructField("mapping_json_str", StringType()),
                StructField("source_sql", StringType()),
                StructField("target_sql", StringType()),
            ]
        )

        # Create DataFrame
        df = session.create_dataframe(rows, schema=schema)

        # Add timestamp and convert JSON string to VARIANT
        df = df.select(
            df["mapping_id"],
            current_timestamp().alias("run_timestamp"),
            parse_json(df["mapping_json_str"]).alias("mapping_json"),
            df["source_sql"],
            df["target_sql"],
        )

        # Write to table
        df.write.mode("append").save_as_table(full_table)

        log.append(f"[STEP 4] {len(data)} row(s) inserted into {full_table}  [OK]")
        log.append(f"         Query table: SELECT * FROM {full_table};")

    # Summary
    log.append("")
    log.append("=" * 42)
    log.append(f"  [OK] PHASE 1 COMPLETE - {len(data)} mapping(s)")
    log.append("=" * 42)
    if WRITE_TO_TABLE:
        log.append(f"  Table: {db_prefix}{OUTPUT_SCHEMA}.{OUTPUT_TABLE}")
    log.append(f"  Warnings: {len(warnings)}")
    log.append(f"  Status: SUCCESS")
    log.append("=" * 42)

    return "\n".join(log)
