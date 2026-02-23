"""
Regenerate Phase1_DataInput.xlsx directly in Snowflake/ETLV3 folder

Run this from your repo root:
  python regenerate_excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Ensure ETLV3 folder exists
os.makedirs("Snowflake/ETLV3", exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "MAPPING"

# Colors
SECTION_HDR = "1F3864"
GOLD = "FFD700"
LIGHT_BLUE = "BDD7EE"
LABEL_BG = "E7E6E6"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"


def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def style_section_header(cell):
    cell.fill = PatternFill("solid", start_color=SECTION_HDR, end_color=SECTION_HDR)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()


def style_label(cell, mandatory=False):
    cell.fill = PatternFill("solid", start_color=LABEL_BG, end_color=LABEL_BG)
    cell.font = Font(name="Arial", bold=True, size=9)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = thin_border()
    if mandatory and cell.value and not cell.value.endswith("(*)"):
        cell.value = f"{cell.value} (*)"


def style_value(cell, mandatory=False):
    bg = GOLD if mandatory else WHITE
    cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
    cell.font = Font(name="Arial", size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border()


def style_col_header(cell, mandatory=False):
    bg = GOLD if mandatory else LIGHT_BLUE
    cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
    cell.font = Font(name="Arial", bold=True, color="000000", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


def style_data(cell, shade=False):
    bg = LIGHT_GREY if shade else WHITE
    cell.fill = PatternFill("solid", start_color=bg, end_color=bg)
    cell.font = Font(name="Arial", size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border()


ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 50
for col in range(3, 20):
    ws.column_dimensions[get_column_letter(col)].width = 15

# MAPPING INFORMATION section
row = 1
ws.merge_cells(f"A{row}:B{row}")
c = ws[f"A{row}"]
c.value = "MAPPING INFORMATION"
style_section_header(c)
ws.row_dimensions[row].height = 25

mapping_fields = [
    ("mapping_id", "MAP_001", True),
    ("db_name", "TEST_1_DB", True),
    ("source_system", "FEED", True),
    ("source_schema", "S_SCHEMA", True),
    ("source_table", "DEMO_TABLE, DEMO_TABLE2", True),
    ("source_table_alias", "T1, T2", False),
    ("source_join", "DEMO_TABLE T1 LEFT JOIN DEMO_TABLE2 T2 ON T1.ID = T2.ID", False),
    ("source_lookup_table", "", False),
    ("source_lookup_join_condition", "", False),
    ("source_filter", "", False),
    ("source_date_filter", "$order_date between STRT_DT and END_DT", False),
    ("source_filter_other", "", False),
    ("dedup_logic", "KEEP_LAST", False),
    ("target_schema", "T_SCHEMA", True),
    ("target_table", "TARGETTABLE1", True),
    ("target_join", "", False),
    ("target_filter", "SRC_SYS_CD = ''FEED''", False),
    (
        "target_date_filter",
        "$order_date between COL_X1.STRT_DT and COL_X1.END_DT",
        False,
    ),
    ("target_filter_other", "", False),
    ("scd2_applicable", "Y", True),
    ("scd2_start_date_column", "STRT_DT", False),
    ("scd2_end_date_column", "END_DT", False),
    ("delete_flag_applicable", "Y", True),
    ("sample_set", "100", False),
    ("tables_group", "1", False),
    ("fk_parent_table", "", False),
    ("custom_1", "", False),
    ("custom_2", "", False),
    ("custom_3", "", False),
    ("custom_4", "", False),
]

row = 2
for label, value, mandatory in mapping_fields:
    ws.row_dimensions[row].height = 20
    label_cell = ws[f"A{row}"]
    value_cell = ws[f"B{row}"]
    label_cell.value = label
    value_cell.value = value if value else None
    style_label(label_cell, mandatory)
    style_value(value_cell, mandatory)
    row += 1

# COLUMN MAPPINGS section
row += 1
ws.row_dimensions[row].height = 10

row += 1
ws.merge_cells(f"A{row}:R{row}")
c = ws[f"A{row}"]
c.value = "COLUMN MAPPINGS"
style_section_header(c)
ws.row_dimensions[row].height = 25

row += 1
ws.merge_cells(f"A{row}:R{row}")
leg = ws[f"A{row}"]
leg.value = "🟡 GOLD = Mandatory (*)     🔵 BLUE = Optional"
leg.font = Font(name="Arial", italic=True, size=9)
leg.alignment = Alignment(horizontal="left", vertical="center")
leg.fill = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
ws.row_dimensions[row].height = 18

row += 1
col_headers = [
    ("source_table", True, 18),
    ("source_column", True, 18),
    ("source_keys", True, 10),
    ("source_data_type", True, 16),
    ("source_length", False, 10),
    ("source_precision", False, 10),
    ("source_default_value", False, 16),
    ("source_column_override", False, 20),
    ("target_column", True, 18),
    ("target_keys", True, 10),
    ("target_data_type", True, 16),
    ("target_ref_data", False, 14),
    ("target_defaults", False, 14),
    ("target_lookup", False, 14),
    ("transformationrule", False, 25),
    ("transformationtype", True, 16),
]

ws.row_dimensions[row].height = 30
for col_idx, (hdr, mandatory, width) in enumerate(col_headers, 1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = f"{hdr} (*)" if mandatory else hdr
    style_col_header(cell, mandatory)
    ws.column_dimensions[get_column_letter(col_idx)].width = width

sample_cols = [
    (
        "DEMO_TABLE",
        "PKCOL1",
        "Y",
        "VARCHAR(100)",
        "",
        "",
        "",
        "",
        "COL_1",
        "Y",
        "VARCHAR(100)",
        "",
        "",
        "",
        "",
        "COPY",
    ),
    (
        "DEMO_TABLE",
        "PKCOL2",
        "Y",
        "VARCHAR(100)",
        "",
        "",
        "",
        "",
        "RLTD_COL_1",
        "Y",
        "VARCHAR(100)",
        "",
        "",
        "",
        "",
        "COPY",
    ),
    (
        "DEMO_TABLE",
        "RELKIND",
        "Y",
        "VARCHAR(100)",
        "",
        "",
        "",
        "",
        "COL_3",
        "Y",
        "VARCHAR(100)",
        "",
        "",
        "",
        "",
        "COPY",
    ),
    (
        "DEMO_TABLE",
        "STRT_DT",
        "N",
        "VARCHAR(100)",
        "",
        "",
        "",
        "",
        "STRT_DT",
        "N",
        "DATE",
        "",
        "",
        "",
        "",
        "COPY",
    ),
    (
        "DEMO_TABLE",
        "END_DT",
        "N",
        "VARCHAR(100)",
        "",
        "",
        "",
        "",
        "END_DT",
        "N",
        "DATE",
        "",
        "",
        "",
        "",
        "COPY",
    ),
    (
        "DEMO_TABLE2",
        "STRT_DT",
        "N",
        "DATE",
        "",
        "",
        "",
        "",
        "BUSN_STRT_DT",
        "N",
        "DATE",
        "",
        "",
        "",
        "",
        "COPY",
    ),
    (
        "DEMO_TABLE2",
        "END_DT",
        "N",
        "DATE",
        "",
        "",
        "",
        "",
        "BUSN_END_DT",
        "N",
        "DATE",
        "",
        "",
        "",
        "",
        "COPY",
    ),
]

row += 1
for r_idx, row_data in enumerate(sample_cols):
    current_row = row + r_idx
    ws.row_dimensions[current_row].height = 18
    for c_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=c_idx, value=val if val else None)
        style_data(cell, shade=(r_idx % 2 == 0))

for r_idx in range(len(sample_cols), len(sample_cols) + 10):
    current_row = row + r_idx
    ws.row_dimensions[current_row].height = 18
    for c_idx in range(1, len(col_headers) + 1):
        cell = ws.cell(row=current_row, column=c_idx, value=None)
        style_data(cell, shade=(r_idx % 2 == 0))

# INSTRUCTIONS sheet
ws2 = wb.create_sheet("INSTRUCTIONS")
ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 70

ws2.merge_cells("A1:B1")
t = ws2["A1"]
t.value = "PHASE 1 — INPUT INSTRUCTIONS"
style_section_header(t)
ws2.row_dimensions[1].height = 25

ws2["A2"].value = "Field"
ws2["B2"].value = "Description"
for c in ["A2", "B2"]:
    cell = ws2[c]
    cell.font = Font(name="Arial", bold=True, size=9)
    cell.fill = PatternFill("solid", start_color=LABEL_BG, end_color=LABEL_BG)
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[2].height = 20

instructions = [
    (
        "Layout",
        "Each mapping has MAPPING INFORMATION (table metadata) followed by COLUMN MAPPINGS. Fields marked (*) are mandatory.",
    ),
    (
        "dedup_logic",
        "How to handle duplicates. Values: KEEP_FIRST, KEEP_LAST (default), REJECT.",
    ),
    (
        "source_lookup_table",
        "Optional reference table to join. Example: 'REF_COUNTRY_CODES'.",
    ),
    (
        "Multiple source tables",
        "If source_table has multiple tables, source_join is required.",
    ),
    ("transformationtype (*)", "COPY = direct copy. SQL = use transformationrule."),
]

row = 3
for label, desc in instructions:
    ws2.row_dimensions[row].height = 24 if desc else 10
    c1 = ws2.cell(row=row, column=1, value=label)
    c2 = ws2.cell(row=row, column=2, value=desc)
    if label:
        c1.font = Font(name="Arial", bold=True if "(*)" in label else False, size=9)
        c1.fill = PatternFill(
            "solid",
            start_color=LABEL_BG if label else WHITE,
            end_color=LABEL_BG if label else WHITE,
        )
    else:
        c1.fill = PatternFill("solid", start_color=WHITE, end_color=WHITE)
    c1.alignment = Alignment(horizontal="left", vertical="center")
    c1.border = thin_border()
    c2.font = Font(name="Arial", size=9)
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border = thin_border()
    row += 1

# Save to Snowflake/ETLV3
output_path = "Snowflake/ETLV3/Phase1_DataInput.xlsx"
wb.save(output_path)
print(f"✅ Excel file created: {output_path}")
print(f"   File size: {os.path.getsize(output_path):,} bytes")
print("")
print("Now run:")
print("  git add Snowflake/ETLV3/")
print("  git commit -m 'Add ETLV3 - Phase 1 Parser v1.0.0'")
print("  git push origin main")
